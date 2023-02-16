from datetime import datetime
from openpyxl import load_workbook

Rut="C:\\Users\\USUARIO\\Desktop\\crud\\crud.xlsx"
Rut=r"C:\Users\USUARIO\Desktop\crud\crud.xlsx"

def leer(ruta:str, extraer:str):
    Archivo_Exccel = load_workbook(ruta)
    Hoja_datos= Archivo_Exccel['Datos del crud']
    Hoja_datos=Hoja_datos['A2':'F'+str(Hoja_datos.max_row)]

    info={}

    for i in Hoja_datos:
        
        if isinstance(i[0].value, int):
            info.setdefault(i[0].value,{'tarea':i[1].value, 'descripcion':i[2].value, 'estado':i[3].value, 'fecha inicio':i[4].value, 'fecha finalizacion':i[5].value})

    if not(extraer=='todo'):
        info=filtrar(info, extraer)

    for i in info:
        print('***  Tarea  **')  
        print('id: '+str(i)+'\n'+'titulo: '+str(info[i]['tarea'])+'\n'+'descripcion: '+str(info[i]['descripcion'])+'\n'+'estado'+str(info[i]['estado']) +'\n'+'fecha inico'+str(info[i]['fecha inicio'])+ '\n'+'fecha finalizacion'+str(info[i]['fecha finalizacion'])             )  
        print()
    return          

def filtrar(info:dict, filtro:str):
    aux={}
    for i in info:
        if info[i]['estado']==filtro:
            aux.setdefault(i, info[i])
    return aux

def actualizar(ruta:str, identificador:int, datos_actualizados:dict):
    Archivo_Exccel=load_workbook(ruta)
    Hoja_datos =Archivo_Exccel['Datos del crud']
    Hoja_datos=Hoja_datos['A2':'F'+str(Hoja_datos.max_row)]
    hoja=Archivo_Exccel.active

    titulo=2
    descripcion=3
    estado=4
    fecha_inicio=5
    fecha_finalizacion=6
    encontro=False
    for i in Hoja_datos:
        if i[0].value==identificador:
            fila=i[0].row
            encontro=True
            for d in datos_actualizados:
                if d=='titulo' and not (datosActualizados[d]==''):
                    hoja.cell(row=fila, column=titulo).value=datosActualizados[d]
                elif d=='descripcion' and not (datosActualizados[d]==''):
                    hoja.cell(row=fila, column=descripcion).value=datosActualizados[d] 
                elif d=='estado' and not (datosActualizados[d]==''):
                    hoja.cell(row=fila, column=estado).value=datosActualizados[d] 
                elif d=='fecha inicio' and not (datosActualizados[d]==''):
                    hoja.cell(row=fila, column=fecha_inicio).value=datosActualizados[d] 
                elif d=='fecha finalizacion' and not (datosActualizados[d]==''):
                    hoja.cell(row=fila, column=fecha_finalizacion).value=datosActualizados[d]  
    Archivo_Exccel.save(ruta)  
    if encontro==False:
        print('Error: No existe una tarea con ese ID')
        print()
    return

def agregar(ruta:int, datos:dict):
    Archivo_Exccel=load_workbook(ruta)
    Hoja_datos=Archivo_Exccel['Datos del crud']
    Hoja_datos=Hoja_datos['A2':'F'+str(Hoja_datos.max_row+1)]                                 
    hoja=Archivo_Exccel.active

    titulo=2
    descripcion=3
    estado=4
    fecha_inicio=5
    fecha_finalizacion=6
    for i in Hoja_datos:
        if not(isinstance(i[0].value, int )):
            identificador=i[0].row
            hoja.cell(row=identificador, column=1).value=identificador-1
            hoja.cell(row=identificador, column=titulo).value=datos['titulo']
            hoja.cell(row=identificador, column=descripcion).value=datos['descripcion']
            hoja.cell(row=identificador, column=estado).value=datos['estado']
            hoja.cell(row=identificador, column=fecha_inicio).value=datos['fecha inicio']
            hoja.cell(row=identificador, column=fecha_finalizacion).value=datos['fecha finalizacion']
            break
    Archivo_Exccel.save(ruta)
    return



def borrar(ruta, identificador):
    Archivo_Exccel= load_workbook(ruta)
    Hoja_datos=Archivo_Exccel['Datos del crud']
    Hoja_datos=Hoja_datos['A2':'F'+str(Hoja_datos.max_row)]
    hoja=Archivo_Exccel.active

    titulo=2
    descripcion=3
    estado=4
    fecha_inicio=5
    fecha_finalizacion=6
    encontro=False
    for i in Hoja_datos:
        if i[0].value==identificador:
            fila=i[0].row
            encontro=True

            hoja.cell(row=fila, column=1).value=""
            hoja.cell(row=fila, column=titulo).value=""
            hoja.cell(row=fila, column=descripcion).value=""
            hoja.cell(row=fila, column=estado).value=""
            hoja.cell(row=fila, column=fecha_inicio).value=""
            hoja.cell(row=fila, column=fecha_finalizacion).value=""
    Archivo_Exccel.save(ruta) 
    if encontro==False:
        print('Error: no exixte una tarea con ese Id')
        print()
    return           
            
Rut="C:\\Users\\USUARIO\\Desktop\\crud\\crud.xlsx"     

datosActualizados={'titulo':'', 'descripcion':'', 'estado':'', 'fecha inico':'', 'fecha finalizacion':''}
while True:
    print('indique la accion que desea realizar: ')
    print('Consultar: 1')
    print('Actualizar: 2')
    print('Crear nueva tarea: 3')
    print('Borrar: 4')
    accion= input('Escriba la opcion: ')
    if not(accion=='1') and not (accion=='2') and not (accion=='3') and not (accion=='4'):
        print('Comando invalido por favor eliga una opcion valida')
    elif accion=='1':
        opc_consulta=''
        print('Indique la tarea que desea consultar: ')
        print('Todas las tareas: 1')
        print('En espera: 2')
        print('En ejecucion: 3')
        print('Por aprobar: 4')
        print('Finalizado: 5')
        opc_consulta= input('Escriba la tarea que desea consultar:')
        if opc_consulta=='1':
            print()
            print()
            print('*** Consultando todas las tareas ***')
            leer(Rut,'todo')
        elif opc_consulta=='2':
            print()
            print()
            print('*** Consultando tareas en espera ***')
            leer(Rut,'En espera')
        elif opc_consulta=='3':
            print()
            print()
            print('*** Consultando tareas en ejecucion ***')
            leer(Rut,'En ejecucion')
        elif opc_consulta=='4':
            print()
            print()
            print('*** Consultando tareas por aprobar ***')
            leer(Rut,'Por aprobar')

        elif opc_consulta=='5':
            print() 
            print()
            print('*** Consultando tareas finalizadas ***')
            leer(Rut,'Finalizada')
    elif accion=='2':
        datosActualizados={'titulo':'', 'descripcion':'', 'estado':'', 'fecha inico':'', 'fecha finalizacion':''}
        print('*** Actualizar tarea ***')
        print()
        id_Actualizar=int(input('Indique el ID de la tarea que desea actualizar: '))
        print()
        print('** Nuevo titulo **')
        print('*** Nota: Si no desea actualizar el titulo solo oprima ENTER ***')
        datosActualizados['titulo']=input('Indique el nuevo titulo de la tarea: ')
        print()
        print('** Nueva descripcion **')
        print('*** Nota: Si no desea actualizar la descripcion solo oprima ENTER ***')
        datosActualizados['descripcion']=input('Indique la nueva descripcion de la tarea')
        print()
        print('** Nuevo estado **')
        print('En espera: 2')
        print('En ejecucion: 3')
        print('Por aprobar: 4')
        print('Finalizada: 5')
        print('*** Nota: si no desea actualizar el estado oprima ENTER')
        estadoNuevo=input('Indique el nuevo estado de la tarea: ')
        if estadoNuevo=='2':
            datosActualizados['estado']='En espera'
        if estadoNuevo=='3':
            datosActualizados['estado']='En ejecucion'
        if estadoNuevo=='4':
            datosActualizados['estado']='En aprobar'
        if estadoNuevo=='5':
            now=datetime.now()
            datosActualizados['estado']='Finalizada'
            datosActualizados['fecha finalizacion']=str(now.day)+'/'+str(now.month)+'/'+str(now.year)

        now=datetime.now()
        datosActualizados['fecha inico']=str(now.day)+'/'+str(now.month)+'/'+str(now.year)
        actualizar(Rut,id_Actualizar, datosActualizados)
        print()
    elif accion=='3':
        datosActualizados={'titulo':'', 'descripcion':'', 'estado':'', 'fecha inico':'', 'fecha finalizacion':''}
        print('** Crear nueva Tarea **')
        
        print()
        print('** titulo **')
        print()
        datosActualizados['titulo']=input('Indique el titulo de la tarea: ')
        print()
        print('** descripcion **')
        datosActualizados['descripcion']= input('Indique la descripcion de la tarea : ')
        print()
        datosActualizados['estado']='En espera'
        now=datetime.now()
        datosActualizados['fecha inicio']=str(now.day)+'/'+str(now.month)+'/'+str(now.year)
        datosActualizados['fecha finalizacion']=''
        agregar(Rut, datosActualizados)
    elif accion=='4':
        print('')
        print('** Eliminar tarea **')
        iden=int(input('Indique el ID de la tarea que desea eliminar'))  
        borrar(Rut,iden)  

                   

    
    
